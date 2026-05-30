from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import numpy as np
import pandas as pd

from .signals import add_business_cycle_labels, add_macro_regime_labels


DEFAULT_BCF_WORKBOOK = "US Business Cycle Model Final v.5.xlsx"
DEFAULT_REGIME_WORKBOOK = "US Regime Cycle Model v.2.xlsx"
DEFAULT_INDEX_WORKBOOK = "/Users/thomasbrenner/Documents/Business Cycle Backtest Index Level no bbg link.xlsx"


@dataclass(frozen=True)
class ReturnSheet:
    name: str
    dataframe: pd.DataFrame
    state_cols: list[str]
    asset_cols: list[str]
    asset_classes: dict[str, str] | None = None


def _clean_name(value: object) -> str:
    text = str(value).strip()
    return " ".join(text.split())


def _dedupe_columns(columns: Iterable[object]) -> list[str]:
    seen: dict[str, int] = {}
    result: list[str] = []
    for col in columns:
        name = _clean_name(col)
        if name in {"None", "nan", ""}:
            name = "unnamed"
        count = seen.get(name, 0)
        seen[name] = count + 1
        result.append(name if count == 0 else f"{name}.{count}")
    return result


def _as_month_end(series: pd.Series) -> pd.Series:
    return pd.to_datetime(series, errors="coerce").dt.to_period("M").dt.to_timestamp("M")


def _norm_asset_name(value: object) -> str:
    return " ".join(str(value).strip().split()).lower()


def _find_col(columns: list[str], contains: str, start: int = 0) -> str:
    needle = contains.lower()
    for col in columns[start:]:
        if needle in col.lower():
            return col
    raise KeyError(f"No column containing {contains!r}")


def load_business_cycle_model(path: str | Path, sheet_name: str = "RoC Model (Base)") -> pd.DataFrame:
    raw = pd.read_excel(path, sheet_name=sheet_name, header=1, engine="openpyxl")
    raw.columns = _dedupe_columns(raw.columns)
    raw = raw.rename(columns={raw.columns[0]: "date"})
    raw["date"] = _as_month_end(raw["date"])
    raw = raw.dropna(subset=["date"]).copy()

    cols = list(raw.columns)
    aggregate = _find_col(cols, "Aggregate Score")
    # The workbook has several "Delta" columns; the 3-month rolling one is explicit.
    delta_3m = _find_col(cols, "3 Month Rolling Delta")
    out = pd.DataFrame(
        {
            "date": raw["date"],
            "aggregate_score": pd.to_numeric(raw[aggregate], errors="coerce"),
            "delta_3m": pd.to_numeric(raw[delta_3m], errors="coerce"),
        }
    )
    for optional in ["Pivot", "RoRo", "Phase"]:
        matches = [c for c in cols if optional.lower() in c.lower()]
        if matches:
            out[optional.lower().replace(" ", "_")] = raw[matches[0]]
    out = out.sort_values("date").reset_index(drop=True)
    return add_business_cycle_labels(out)


def load_macro_regime(path: str | Path, sheet_name: str = "Monthly Regime") -> pd.DataFrame:
    raw = pd.read_excel(path, sheet_name=sheet_name, header=0, engine="openpyxl")
    raw.columns = _dedupe_columns(raw.columns)
    raw = raw.rename(columns={raw.columns[0]: "date"})
    raw["date"] = _as_month_end(raw["date"])
    raw = raw.dropna(subset=["date"]).copy()
    cols = list(raw.columns)
    growth = _find_col(cols, "3-month CLI")
    inflation = _find_col(cols, "3-month CPI")
    regime_col = _find_col(cols, "Regime")
    out = pd.DataFrame(
        {
            "date": raw["date"],
            "growth_delta_3m": pd.to_numeric(raw[growth], errors="coerce"),
            "inflation_delta_3m": pd.to_numeric(raw[inflation], errors="coerce"),
            "regime": raw[regime_col],
        }
    )
    out = out.sort_values("date").reset_index(drop=True)
    return add_macro_regime_labels(out)


def _normalize_return_units(df: pd.DataFrame, asset_cols: list[str]) -> pd.DataFrame:
    out = df.copy()
    for col in asset_cols:
        values = pd.to_numeric(out[col], errors="coerce")
        max_abs = values.abs().quantile(0.99)
        if pd.notna(max_abs) and max_abs > 2:
            values = values / 100.0
        out[col] = values.replace([np.inf, -np.inf], np.nan)
    return out


def load_state_return_sheet(path: str | Path, sheet_name: str) -> ReturnSheet:
    df = pd.read_excel(path, sheet_name=sheet_name, header=0, engine="openpyxl")
    df.columns = _dedupe_columns(df.columns)
    df = df.rename(columns={df.columns[0]: "date"})
    df["date"] = _as_month_end(df["date"])
    df = df.dropna(subset=["date"]).copy()

    state_candidates = [
        col
        for col in df.columns
        if col.lower() in {"phase", "phase t", "phase t-1", "phase (t-1)", "regime", "regime t-1"}
        or col.lower().startswith("phase ")
        or col.lower().startswith("regime ")
    ]
    state_cols = [c for c in state_candidates if df[c].notna().any()]
    if "Phase" in df.columns and "Regime" in df.columns:
        state_cols = list(dict.fromkeys(["Phase", "Regime", *state_cols]))
    if "Phase T" in df.columns and "Regime" in df.columns:
        state_cols = list(dict.fromkeys(["Phase T", "Regime", *state_cols]))

    first_asset_idx = 1
    if state_cols:
        first_asset_idx = max(df.columns.get_loc(c) for c in state_cols) + 1
    raw_asset_cols = []
    for col in list(df.columns)[first_asset_idx:]:
        low = col.lower()
        if (
            low.startswith("unnamed")
            or "summary" in low
            or "t-bill rate" in low
            or "rfr monthly" in low
            or low == "enora"
        ):
            continue
        numeric = pd.to_numeric(df[col], errors="coerce")
        if numeric.notna().sum() >= 6:
            raw_asset_cols.append(col)

    er_cols = [col for col in raw_asset_cols if "(er)" in col.lower()]
    asset_cols = er_cols if er_cols else raw_asset_cols

    df = _normalize_return_units(df, asset_cols)
    if {"Phase", "Regime"}.issubset(df.columns):
        df["Phase-Regime"] = df["Phase"].astype(str) + "-" + df["Regime"].astype(str)
        state_cols.append("Phase-Regime")
    if {"Phase T", "Regime"}.issubset(df.columns):
        df["Phase-Regime"] = df["Phase T"].astype(str) + "-" + df["Regime"].astype(str)
        state_cols.append("Phase-Regime")

    state_cols = [c for c in dict.fromkeys(state_cols) if c in df.columns]
    return ReturnSheet(sheet_name, df.sort_values("date").reset_index(drop=True), state_cols, asset_cols)


def load_index_asset_classes(path: str | Path) -> dict[str, str]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    ws = workbook["Raw Data"]
    headers = {str(ws.cell(1, col).value).strip(): col for col in range(1, ws.max_column + 1)}
    name_col = headers.get("Name")
    class_col = headers.get("Global Broad Category Group")
    if not name_col or not class_col:
        raise ValueError("Raw Data sheet is missing Name or Global Broad Category Group.")
    mapping: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        asset = _clean_name(row[name_col - 1])
        asset_class = _clean_name(row[class_col - 1])
        if asset and asset_class and asset_class.lower() != "nan":
            mapping[_norm_asset_name(asset)] = asset_class
    return mapping


def load_index_level_returns(path: str | Path) -> ReturnSheet:
    """Load index-level returns using Index Monthly Prices as source of truth.

    Friendly index names are on row 1, Bloomberg tickers are on row 2, and prices
    begin on row 3. Returns are calculated locally to avoid depending on Excel
    formula caches.
    """
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    ws = workbook["Index Monthly Prices"]
    first_two_rows = list(ws.iter_rows(min_row=1, max_row=2, values_only=True))
    asset_names = [_clean_name(value) for value in first_two_rows[0][1:]]

    rows: list[list[object]] = []
    for values in ws.iter_rows(min_row=3, values_only=True):
        date = values[0]
        if date is None:
            continue
        rows.append([date, *values[1 : len(asset_names) + 1]])
    prices = pd.DataFrame(rows, columns=["date", *asset_names])
    prices["date"] = _as_month_end(prices["date"])
    prices = prices.dropna(subset=["date"]).sort_values("date").reset_index(drop=True)

    asset_cols = [col for col in prices.columns if col != "date"]
    for col in asset_cols:
        prices[col] = pd.to_numeric(prices[col], errors="coerce")
    returns = prices.copy()
    returns[asset_cols] = prices[asset_cols].pct_change()
    returns = returns.iloc[1:].reset_index(drop=True)

    state_rows: list[list[object]] = []
    ws_states = workbook["Monthly Excess Returns"]
    for values in ws_states.iter_rows(min_row=2, max_col=3, values_only=True):
        date = values[0]
        if date is None:
            continue
        state_rows.append([date, values[1], values[2]])
    states = pd.DataFrame(state_rows, columns=["date", "Phase t-1", "Regime t-1"])
    states["date"] = _as_month_end(states["date"])
    state_cols = [col for col in ["Phase t-1", "Regime t-1"] if col in states.columns]
    states = states[["date", *state_cols]].dropna(subset=["date"])

    df = returns.merge(states, on="date", how="left")
    if {"Phase t-1", "Regime t-1"}.issubset(df.columns):
        df["Phase-Regime t-1"] = df["Phase t-1"].astype(str) + "-" + df["Regime t-1"].astype(str)
        state_cols.append("Phase-Regime t-1")

    class_lookup = load_index_asset_classes(path)
    asset_classes = {
        asset: class_lookup.get(_norm_asset_name(asset), "Unclassified")
        for asset in asset_cols
    }
    return ReturnSheet("Index Level", df, state_cols, asset_cols, asset_classes)


def build_asset_class_returns(return_sheet: ReturnSheet) -> ReturnSheet:
    if not return_sheet.asset_classes:
        raise ValueError("ReturnSheet has no asset class mapping.")
    df = return_sheet.dataframe[["date", *return_sheet.state_cols]].copy()
    class_to_assets: dict[str, list[str]] = {}
    for asset in return_sheet.asset_cols:
        asset_class = return_sheet.asset_classes.get(asset, "Unclassified")
        class_to_assets.setdefault(asset_class, []).append(asset)
    class_cols = []
    for asset_class, assets in sorted(class_to_assets.items()):
        df[asset_class] = return_sheet.dataframe[assets].mean(axis=1, skipna=True)
        class_cols.append(asset_class)
    return ReturnSheet("Asset Class", df, return_sheet.state_cols, class_cols, {c: c for c in class_cols})
